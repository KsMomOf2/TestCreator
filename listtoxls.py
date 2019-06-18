import openpyxl


class ListToXLS:


	def __init__(self, filename, headers, questions):
		self.filename = filename[:len(filename)-4] + 'xlsx'
		self.headers = headers
		self.questions = questions

		wb, sheet = self.create_sheet()
		self.write_headers(sheet)
		self.write_questions(sheet)
		self.save_sheet(wb, sheet)

	def create_sheet(self):
		wb = openpyxl.Workbook()
		sheet = wb.active
		return wb, sheet

	def write_headers(self, sheet):
		h = self.headers
		sheet.cell(row = 1, column=1).value = h.section
		sheet.cell(row = 2, column=1).value = h.test_name
		sheet.cell(row = 3, column=1).value = h.instructions
		# write the column headers
		sheet.cell(row = 4, column=1).value = 'Num'
		sheet.cell(row = 4, column = 2).value = 'Question'
		sheet.cell(row = 4, column = 3).value = 'Answer'
		sheet.cell(row = 4, column = 4).value = 'Choices'
		# name the sheet
		sheet.title = h.test_name

	def write_questions(self, sheet):
		rowNum = 5
		for q in self.questions:
			rowNum = rowNum + 1
			sheet.cell(row = rowNum, column=1).value = q.number
			sheet.cell(row = rowNum, column=2).value = q.question
			sheet.cell(row = rowNum, column=3).value = q.answer
			colNum = 4
			for c in q.choices:
				sheet.cell(row = rowNum, column=colNum).value = c
				colNum = colNum + 1

	def save_sheet(self, wb, sheet):
		wb.save(self.filename)

